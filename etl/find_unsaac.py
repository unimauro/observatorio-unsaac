import openpyxl, unicodedata, glob, os
from collections import Counter
CONOSCE=os.environ.get('CONOSCE_DIR','/Users/unimauro/Documents/Repos/observatorio-sanmarcos/etl/conosce')
def norm(s):
    if s is None: return ''
    s=''.join(c for c in unicodedata.normalize('NFD',str(s)) if unicodedata.category(c)!='Mn')
    return s.upper()
for Y in (2023,2024,2025):
    f=f"{CONOSCE}/CONOSCE_ADJUDICACIONES{Y}_0.xlsx"
    wb=openpyxl.load_workbook(f, read_only=True)
    ws=wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True); next(rows)
    hits=Counter()
    for r in rows:
        ent=norm(r[2])
        if 'SAN ANTONIO ABAD' in ent or ('CUSCO' in ent and 'UNIVERSIDAD' in ent):
            hits[(r[0], r[1], r[2])]+=1
    print(f"--- {Y} ---")
    for k,v in sorted(hits.items(), key=lambda x:-x[1]):
        print(v, k)
    wb.close()
